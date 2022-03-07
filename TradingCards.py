import openpyxl
class Card:
    """
    Card has the information of the cards in a deck 
    """
    def __init__(self, theName, theType, theHP, theMoves, isShiny):
        """
        __init__ Function has the details of the Card
        """
        self.name = str(theName)
        self.type = theType
        self.HP = theHP
        self.moves = theMoves
        self.shiny = isShiny
    def __str__(self):
        """
        prints the selected 
        """
        print("\nName: {self.name}".format(self=self))
        print("Type: {self.type}".format(self=self))
        print("Maximum HP: {self.HP}".format(self=self))
        print("Moves:")
        for i in self.moves:
            print(i)
        if(self.shiny == 0):
            return "Shiny Status: Not Shiny"
        else:     
            return "Shiny Status: Shiny!"
    def save(self):
        """
        Saves the card details
        """
        card = [self.name, self.type, self.HP, self.shiny]
        for i in self.moves:
            for x in i:
                card.append(x)
        return card
class Deck: 
    orderofDeck = []
    def __init__(self):
        """
        has a list which holds cards in the deck
        """
        self.deck = []
    def inputFromFile(self, fileName):
        """
        reads a spreadsheet file containing card details
        """
        self.fileName = str(fileName)
        workBook = openpyxl.load_workbook(self.fileName)
        activeSheet = workBook.active
        for row in activeSheet.iter_rows(min_row = 2):            
            cards = []
            for cell in row:             
                cards.append(cell.value)
            if(len(cards) < 6):
                continue
            else:
                if(cards[0] is None):
                    continue
                else:
                    name = row[0].value
                    name = str(name)                
                    Type = row[1].value
                    if(Type != "Magi" and Type != "Water" and Type != "Fire" and Type != "Earth" and Type != "Air" and Type != "Astral"):
                        print("Invalid Card Type.", name, " not registered.")
                        print("Valid card types: 'Magi', 'Water', 'Fire', 'Earth', 'Air' or 'Astral'")
                        pass
                    else:
                        hp = row[2].value
                        try:
                            float(hp)
                        except TypeError:
                            print("HP entered not valid.", name, "not registered.")
                            pass
                        else:
                            if(type(hp) != int):
                                hp = int(hp)
                                print("HP should be entered as an integer")
                            shiny = row[3].value
                            if(shiny != 0 and shiny !=1):
                                print("Shiny status should be either '0' or '1'")
                                shiny = 0
                            length = len(cards)
                            moves = []
                            movesOut = []
                            for i in range(4, length, 2):
                                move = []
                                if(row[i].value is None):
                                    pass
                                else:
                                    move.append(row[i].value)
                                    if(row[i+1].value is None):
                                        move.append(0)
                                        moves.append(move)
                                    else:
                                        move.append(row[i+1].value)
                                        moves.append(move)
                            if((len(moves)) > 5):
                                print(name, " has more than five moves.")
                                del moves[5:]
                            for i in moves:
                                if(len(i) != 2):
                                    print("Move needs to be entered as a list")
                                    print(i, " removed as a move for", name)
                                    movesOut.append(moves.index(i))
                                else:    
                                    if (isinstance(i[0], str) == False):
                                        i[0] = str(i[0])
                                    if (isinstance(i[1], int) == False):
                                        try:
                                            int(i[1])
                                        except ValueError:
                                            print("invalid damage type ")
                                            i[1] = 0
                                        i[1] = int(i[1])
                            movesOut.reverse()
                            for i in movesOut:
                                del moves[i] 
                            self.deck.append(Card(name, Type, hp, moves, shiny))
                            totalMoves = len(moves)
                            totalDamage = 0
                            for i in moves:
                                totalDamage += i[1]
                            avgDamage = (totalDamage/totalMoves)   
                            Deck.orderofDeck.append([name, Type, hp, avgDamage, shiny])
        self.numOfCards = len(self.deck)
    def __str__(self):
        """
        gives the Card details in the deck
        """
        self.numOfCards = len(self.deck)
        for i in range(0, self.numOfCards):
            print(Deck.getDeck(self)[i])
    def addCard(self, theCard):
        """
        addCard Function adds the card to the deck
        """
        if(isinstance(theCard, list)):
            if(len(theCard) == 5):
                name = theCard[0]
                Type = theCard[1]
                hp = theCard[2]
                moves = theCard[3]
                shiny = theCard[4]
            else:
                print("Card does not have all requirements")
                return
        else:
            print("Provide all the details of Card")
            return  
        Name = str(name)                
        if(Type != "Magi" and Type != "Water" and Type != "Fire" and Type != "Earth" and Type != "Air" and Type != "Astral"):
            print("Card Type Invalid.")
            return
        else:
            try:
                float(hp)
            except ValueError:
                print("HP not valid")
                return
            else:
                if(type(hp) != int):
                    hp = int(hp)
                    print("HP should be an integer")
                if(shiny != 1 and shiny !=0 ):
                    print("if card is shiny enter '1' and if not '0'")
                    shiny = 0
                else:
                    for i in moves:
                        if((isinstance(i, list)) == False):
                            print("Move needs to a list")
                            return
                if((len(moves)) > 5):
                    print(name, " has more than five moves")
                    del moves[5:]
                movesOut = []
                for i in moves:
                    if(len(i) != 2):
                        print("Move needs to be a list")
                        movesOut.append(moves.index(i))
                    else:    
                        if (isinstance(i[0], str) == False):
                            i[0] = str(i[0])
                        if (isinstance(i[1], int) == False):
                            try:
                                int(i[1])
                            except ValueError:
                                print("type for move ", i[0], " invalid")
                                i[1] = 0
                            i[1] = int(i[1])
                movesOut.reverse()
                for i in movesOut:
                    del moves[i] 
                self.deck.append(Card(name, Type, hp, moves, shiny))
                totalMoves = len(moves)
                totalDamage = 0
                for i in moves:
                    totalDamage += i[1]
                avgDamage = (totalDamage / totalMoves)   
                Deck.orderofDeck.append([name, Type, hp, avgDamage, shiny])     
    def rmCard(self, theCard):
        """
        removes selected card from the deck
        """
        try:
            index = Deck.orderofDeck.index(theCard)
        except ValueError:
            print(theCard, "was not found in deck")
        else:
            index = Deck.orderofDeck.index(theCard)
            del self.deck[index]
            del Deck.orderofDeck[index]
    def getMostPowerful(self):
        """
        gives the most powerful card 
        """
        Powerful = 0
        count = 0
        for i in Deck.orderofDeck:
            if(i[3] > Powerful):
                Powerful = i[3]
                index = count
            count += 1
            print("Most Powerful Card: ",Deck.getDeck(self)[index])
    def getAverageDamage(self):
        """
        finds average damage of cards in deck
        """
        totalavgDam = 0
        self.numOfCards = len(self.deck)
        for i in Deck.orderofDeck:
            totalavgDam += i[3]
        totalavgDam = round((totalavgDam/self.numOfCards), 1)
        print("Total Average Damage: ", totalavgDam)
    def viewAllCards(self):
        """
        shows all the cards in the deck   
        """
        Deck.__str__(self)
    def viewAllShinyCards(self):
        """
        displays all the shiny cards in the deck  
        """
        index = []
        for i in Deck.orderofDeck:
            if(i[4] == 1):
                index.append(Deck.orderofDeck.index(i))
        for i in index:
            print(Deck.getDeck(self)[i])
    def viewAllByType(self, theType):
        """
        Displays all cards of selected type 
        """
        self.type = theType
        if(self.type != "Magi" and self.type != "Water" and self.type != "Fire" and self.type != "Earth" and self.type != "Air" and self.type != "Astral"):
            print("Card Type is not valid")
        else:
            index = []
            for i in Deck.orderofDeck:
                if(i[1] == self.type):
                    index.append(Deck.orderofDeck.index(i))
            for i in index:
                print(Deck.getDeck(self)[i])
    def getCards(self):
        """
        returns the selected card from the deck
        """
        return Deck.orderofDeck
    def getDeck(self):
        """
        returns the details of cards in deck
        """
        return self.deck
    def saveToFile(self,fileName):
        """
        creates a new spreadsheet and saves the details of cards 
        """
        fileName = str(fileName)
        if((fileName.endswith(".xlsx")) == False):
            fileName += ".xlsx"
        self.numOfCards = len(self.deck)
        record = []
        for i in Deck.getDeck(self):
            record.append(Card.save(i))
        newBook = openpyxl.Workbook()
        newSheet = newBook.active
        firstRow = ("Name", "Type", "HP", "Shiny", "Move Name 1", "Damage 1", "Move Name 2", "Damage 2", "Move Name 3", "Damage 3", "Move Name 4", "Damage 4", "Move Name 5", "Damage 5")
        newSheet.append(firstRow)
        for i in record:
            newSheet.append(i)
        newBook.save(fileName)
def main():
    """
    Main will validate the file name
    """
    if __name__ == "__main__":
        main()
 

