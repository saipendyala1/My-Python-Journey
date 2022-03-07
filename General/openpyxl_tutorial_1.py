# reading an Excel Spreadsheet

#openpyxl is a Python library to read/write Excel xlsx/xlsm/xltx/xltm files.
from openpyxl import load_workbook

# reading an excel spreadsheet / workbook

# use load_workbook() method for loading excel spreadsheet
workbook = load_workbook(filename="sample.xlsx")

# use .sheetnames to see all sheets
print('workbook.sheetnames: ', workbook.sheetnames)
print()

# .active selects sheet 1 automatically
active_sheet = workbook.active
print('active_sheet: ', active_sheet)
print()

sheet_title = active_sheet.title
print('sheet_title: ', sheet_title)
print()

'''

# Additional Reading options

# read_only loads a spreadsheet in read-only mode 
# allowing you to open very large Excel files.
# workbook = load_workbook(filename="sample.xlsx", read_only=True)


# data_only ignores loading formulas and instead 
# loads only the resulting values.
# workbook = load_workbook(filename="sample.xlsx", data_only=True)


# prints cell object
A1 = active_sheet["A1"]
print('A1: ', A1)
print()

# prints cell value
A1_value = active_sheet["A1"].value
print('A1_value: ', A1_value)
print()

# prints cell value
F2 = active_sheet["F2"].value
print('F2: ', F2)
print()

# note spreadsheets used one-indexed notation
# prints cell value
F5 = active_sheet.cell(row=5, column=6)
print('F5: ', F5)
print()

# prints cell value
F5 = active_sheet.cell(row=5, column=6).value
print('F5: ', F5)
print()

A1_to_C2 = active_sheet["A1:C2"]
print('A1_to_C2: ', A1_to_C2)
print()

# Python generators - to go through the data
# .iter_rows() and .iter_cols()

# Both methods recieve the following arguments:
# min_row, max_row, min_col, max_col

for row in active_sheet.iter_rows(min_row=1,
                           max_row=2,
                           min_col=1,
                           max_col=3):
    print(row)
print()

for column in active_sheet.iter_cols(min_row=1,
                              max_row=2,
                              min_col=1,
                              max_col=3):
    print(column)
print()

 # notice that iterating through the rows using .iter_rows(), 
 # you get one tuple element per row 
 # when using .iter_cols() and iterating through columns, 
 # you’ll get one tuple per column

# To get values
for value in active_sheet.iter_rows(min_row=1,
                             max_row=2,
                             min_col=1,
                             max_col=3,
                             values_only=True):
    print(value)
print()

# manipulating data using python structures

# get headers
for value in active_sheet.iter_rows (min_row=1, max_row=1, values_only=True):
    print(value)
print()

# let us say we are interested in:
# product_id, product_parent, product_title, product_category

# add requires data to a dictionary


for value in active_sheet.iter_rows (min_row=2, max_row=4, min_col=4, max_col=7, values_only=True):
    print(value)
print()


products = {}

# Using the values_only because you want to return the cells' values
for row in active_sheet.iter_rows(min_row=2,
                           max_row=4,
                           min_col=4,
                           max_col=7,
                           values_only=True):
    product_id = row[0]
    product = {
        "parent": row[1],
        "title": row[2],
        "category": row[3]
    }
    products[product_id] = product

print(products)
print()

products = {}

# Using the values_only because you want to return the cells' values
for row in active_sheet.iter_rows(min_col=4,
                           max_col=7,
                           values_only=True):
    product_id = row[0]
    product = {
        "parent": row[1],
        "title": row[2],
        "category": row[3]
    }
    products[product_id] = product

print('Length of prducts dictionary: ', len(products))
print()

# convert data into Python classes

# data class is a class typically coontaining mainly data
# created using the @dataclass decorator,

from dataclasses import dataclass 

@dataclass
class Employee:
    name: str
    salary: int

employee1 = Employee("Bob", 2000)
print(employee1.name)
print(employee1.salary)
print()


for value in active_sheet.iter_rows(min_row=1,
                             max_row=1,
                             values_only=True):
    print(value)
print()


# import methods
import mapping
import classes

#from classes import Product, Review
#from mapping import PRODUCT_ID, PRODUCT_PARENT, PRODUCT_TITLE, \
#    PRODUCT_CATEGORY, REVIEW_DATE, REVIEW_ID, REVIEW_CUSTOMER, \
#    REVIEW_STARS, REVIEW_HEADLINE, REVIEW_BODY

#product = Product()
#print(PRODUCT_CATEGORY)


products = []
reviews = []

for row in active_sheet.iter_rows(min_row=2, max_row=4, values_only=True):
    product = classes.Product(id=row[mapping.PRODUCT_ID],
                              parent=row[mapping.PRODUCT_PARENT],
                              title=row[mapping.PRODUCT_TITLE],
                              category=row[mapping.PRODUCT_CATEGORY])
    products.append(product)

    review = classes.Review(id=row[mapping.REVIEW_ID],
                            customer_id=row[mapping.REVIEW_CUSTOMER],
                            stars=row[mapping.REVIEW_STARS],
                            headline=row[mapping.REVIEW_HEADLINE],
                            body=row[mapping.REVIEW_BODY])

    reviews.append(review)


print(products[0])
print(products[1])
print()

print(reviews[0])
print(reviews[1])
print()

print(type(products[0]))
print(type(reviews[0]))

'''



'''
# creating an excel sheet
from openpyxl import Workbook

workbook = Workbook()
sheet = workbook.active

sheet["A1"] = "hello"
sheet["B1"] = "world!"

workbook.save(filename="hello_world.xlsx")

'''

