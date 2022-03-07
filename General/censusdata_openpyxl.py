#openpyxl is a Python library to read/write Excel xlsx/xlsm/xltx/xltm files.
from openpyxl import load_workbook
from collections import defaultdict

# reading an excel spreadsheet / workbook

# use load_workbook() method for loading excel spreadsheet
workbook = load_workbook(filename="censuspopdata.xlsx")

# use .sheetnames to see all sheets
print('workbook.sheetnames: ', workbook.sheetnames)
print()

# .active selects sheet 1 automatically
active_sheet = workbook.active
print('active_sheet: ', active_sheet)
print()


for row in active_sheet.iter_rows(min_row=1,
                           max_row=2,
                           min_col=1,
                           max_col=4,
                           values_only=True):
    print(row)
print()

county_dict = defaultdict(list)

for row in active_sheet.iter_rows(min_row=2,
                           max_row=72865,
                           min_col=1,
                           max_col=4,
                           values_only=True):
    county_name = row[2]
    censusdata = {
        "censustract": row[0],
        "population": row[3],
    }

    county_dict[county_name].append(censusdata)

#print(county_dict)

print()
popCountList = []
for key, values in sorted(county_dict.items()):
    count = 0
    for val in values:
        count = count + val['population']
    tmp = []
    tmp.append(key)
    tmp.append(str(count))
    popCountList.append(tmp)

# writing list to text file
f = open("population_count.txt", "w")
for item in popCountList:
    f.write(' '.join(item)+'\n')

f.close()
